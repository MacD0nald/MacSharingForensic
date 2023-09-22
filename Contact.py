import os
import plistlib
from datetime import datetime
import shutil
import openpyxl
import getpass  # getpass 모듈 추가

def directory():

    # 사용자 이름 가져오기
    current_user = getpass.getuser()

    # 경로에 사용자 이름 추가
    source_folder = os.path.expanduser(f"~/Library/Application Support/AddressBook/Sources")

    current_script_directory = os.path.dirname(__file__)

    # 대상 폴더 경로 (현재 스크립트 파일이 있는 폴더로 상대 경로 지정)
    destination_folder = os.path.join(current_script_directory, "destination_folder")

    # 대상 폴더가 없다면 생성
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    # 원본 폴더에서 폴더 목록 가져오기
    files = os.listdir(source_folder)
    files.remove('.DS_Store')
    
    for i in files:
        source_folder2 = os.path.join(source_folder, i)
        source_folder2 = os.path.join(source_folder2, "Metadata")
        files = os.listdir(source_folder2)

        # 파일을 순회하면서 복사 및 이름 변경
        for i, filename in enumerate(files):
            print(filename)
            if filename.endswith('.abcdp'):
                print(filename)
                source_file_path = os.path.join(source_folder2, filename)
                # 복사한 파일을 대상 폴더에 저장
                new_filename = f'{i+1}.abcdp'
                destination_file_path = os.path.join(destination_folder, new_filename)
                shutil.copy(source_file_path, destination_file_path)
                # 확장자를 변경하기 위해 파일 이동
                new_extension = '.plist'
                renamed_file_path = os.path.join(destination_folder, f'{i+1}{new_extension}')
                os.rename(destination_file_path, renamed_file_path)
    return destination_folder

def read_bplist(file_path):
    with open(file_path, 'rb') as fp:
        plist_data = plistlib.load(fp)
    return plist_data

def sel_bplist(plist_data):
    desired_keys = ['First', 'Last', 'Organization', 'Creation', 'Modification']
    filtered_dict = {}

    for key in desired_keys:
        value = plist_data.get(key, '')
        filtered_dict[key] = value

    Phone_dict = plist_data.get('Phone', {})
    
    values = Phone_dict.get('values', [])
    
    filtered_dict['phone number'] = values[0] if values else None
    
    for key in ['Creation', 'Modification']:
        if key in filtered_dict:
            creation_datetime = filtered_dict[key]
            formatted_creation = creation_datetime.strftime('%Y년 %m월 %d일 %H시 %M분 %S.%f초')
            filtered_dict[key] = formatted_creation
    return filtered_dict

    

if __name__ == "__main__":
    destination_folder = directory()
    all_filtered_data = []

    for filename in os.listdir(destination_folder):
        if filename.endswith(".plist"):
            file_path = os.path.join(destination_folder, filename)
            
            plist_data = read_bplist(file_path)
            
            filtered_data = sel_bplist(plist_data)
            all_filtered_data.append(filtered_data)

    # 엑셀 워크북 생성
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 열 제목 입력
    columns = list(all_filtered_data[0].keys())
    for col_num, column_title in enumerate(columns, 1):
        sheet.cell(row=1, column=col_num, value=column_title)

    # 데이터 입력
    for row_num, data in enumerate(all_filtered_data, 2):
        for col_num, cell_value in enumerate(data.values(), 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # 엑셀 파일 저장
    current_script_directory = os.path.dirname(__file__)
    excel_file_path = os.path.join(current_script_directory, 'output.xlsx')
    workbook.save(excel_file_path)

    print(f"엑셀 파일이 {excel_file_path} 경로에 생성되었습니다.")
